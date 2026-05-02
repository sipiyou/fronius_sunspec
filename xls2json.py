#!/usr/bin/env python3
"""
xls2json.py – Fronius Modbus Register Map XLSX → Edomi-Importer JSON
(c) Nima Ghassemi Nejad

Jede XLSX-Datei erzeugt ein eigenes JSON.
Blockstartadressen werden automatisch aus dem "Complete Map"-Sheet ermittelt.
Blattnamen werden als Gruppennamen übernommen.

Abhängigkeit:  pip install openpyxl
Aufruf:
  python3 xls2json.py                          # alle XLSX im extracted_xls/-Ordner
  python3 xls2json.py file1.xlsx file2.xlsx    # nur bestimmte Dateien
  python3 xls2json.py --outdir /tmp/out        # Ausgabe in anderen Ordner
"""

import argparse
import json
import os
import re
import sys
import openpyxl

XLSX_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "extracted_xls")
OUT_DIR  = os.path.dirname(os.path.abspath(__file__))

SKIP_SHEETS = {"complete map", "end block"}

TYPE_MAP = {
    "string": "string32",
    "count":  "uint16",
    "pad":    "uint16",
}


def sheet_to_group_name(name: str) -> str:
    """'inverter (10x)' → 'inverter_10x'"""
    name = name.strip()
    name = re.sub(r'[\s\(\)]+', '_', name)
    name = re.sub(r'_+', '_', name)
    return name.strip('_')


def xlsx_to_device_name(path: str) -> str:
    """'Gen24_Primo_Symo_Inverter_Register_Map_Int&SF_storage_ROW.xlsx' → 'Gen24_Primo_Symo_Inverter_Int+SF_storage_ROW'"""
    base = os.path.splitext(os.path.basename(path))[0]
    # 'Register_Map_' entfernen
    base = re.sub(r'_?Register_Map_?', '_', base)
    base = re.sub(r'_+', '_', base).strip('_')
    return base


def xlsx_to_json_name(path: str) -> str:
    base = os.path.splitext(os.path.basename(path))[0]
    base = re.sub(r'[^A-Za-z0-9_&+.-]', '_', base)
    return base + ".json"


def normalize_type(t: str) -> str:
    t = (t or "").strip().lower()
    return TYPE_MAP.get(t, t)


def get_block_starts(ws) -> list:
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next((i for i, r in enumerate(rows) if r[0] == 'Start'), None)
    if hdr_idx is None:
        return []
    starts = []
    for row in rows[hdr_idx + 1:]:
        if row[0] is None:
            continue
        try:
            addr = int(row[0])
        except (ValueError, TypeError):
            continue
        name = (row[5] or "").strip()
        desc = (row[6] or "").strip().lower()
        if name == 'ID' and 'end block' not in desc:
            starts.append(addr)
    return starts


def parse_sheet(ws, block_start: int) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx = next((i for i, r in enumerate(rows) if r[0] == 'Start'), None)
    if hdr_idx is None:
        return {}

    result = {}
    line = 0
    for row in rows[hdr_idx + 1:]:
        if row[0] is None:
            continue
        try:
            rel_start = int(row[0])
        except (ValueError, TypeError):
            continue

        line += 1

        def col(i):
            v = row[i] if i < len(row) else None
            return str(v).replace("\n", " ").replace("\r", " ").strip() if v is not None else ""

        try:
            size = int(row[2])
        except (ValueError, TypeError):
            size = 1

        range_val    = col(10)
        not_supported = 1 if 'not supported' in range_val.lower() else 0

        result[line + 1] = {
            "start":        rel_start + block_start - 1,
            "size":         size,
            "rw":           col(3),
            "function":     col(4),
            "name":         col(5),
            "desc":         col(6),
            "type":         normalize_type(col(7)),
            "unit":         col(8),
            "scaleFactor":  col(9),
            "range":        range_val,
            "notsupported": not_supported,
        }
    return result


def process_xlsx(path: str) -> dict | None:
    if not os.path.exists(path):
        print(f"  FEHLER: Datei nicht gefunden: {path}", file=sys.stderr)
        return None

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        print(f"  FEHLER beim Öffnen {path}: {e}", file=sys.stderr)
        return None

    if "Complete Map" not in wb.sheetnames:
        print(f"  WARNUNG: Kein 'Complete Map'-Sheet – übersprungen: {os.path.basename(path)}", file=sys.stderr)
        return None

    block_starts = get_block_starts(wb["Complete Map"])
    data_sheets  = [s for s in wb.sheetnames if s.strip().lower() not in SKIP_SHEETS]

    if len(data_sheets) != len(block_starts):
        print(f"  WARNUNG: {len(data_sheets)} Sheets, {len(block_starts)} Blöcke – "
              f"prüfe SKIP_SHEETS ({os.path.basename(path)})", file=sys.stderr)

    elements = []
    for sheet_name, block_start in zip(data_sheets, block_starts):
        group_name = sheet_to_group_name(sheet_name)
        regs       = parse_sheet(wb[sheet_name], block_start)
        n_skip     = sum(1 for r in regs.values() if r["notsupported"])
        print(f"    {sheet_name:35s} → {group_name:28s}  start={block_start}  "
              f"regs={len(regs)}  not_supported={n_skip}")
        elements.append({
            "firstID":  block_start,
            "group":    group_name,
            "elements": regs,
        })

    return {
        "device":   xlsx_to_device_name(path),
        "protocol": "TCP",
        "elements": elements,
    }


def main():
    parser = argparse.ArgumentParser(description="Fronius XLSX → Edomi Modbus JSON")
    parser.add_argument("files", nargs="*",
                        help="XLSX-Dateien (Default: alle im extracted_xls/-Ordner)")
    parser.add_argument("--outdir",  default=OUT_DIR,         help=f"Ausgabeverzeichnis (Default: {OUT_DIR})")
    args = parser.parse_args()

    if args.files:
        paths = [f if os.path.isabs(f) else os.path.join(XLSX_DIR, f) for f in args.files]
    else:
        paths = sorted(
            os.path.join(XLSX_DIR, f)
            for f in os.listdir(XLSX_DIR)
            if f.lower().endswith(".xlsx")
        )

    os.makedirs(args.outdir, exist_ok=True)

    ok = 0
    for path in paths:
        print(f"\n{os.path.basename(path)}")
        device = process_xlsx(path)
        if device is None:
            continue

        out_name = os.path.join(args.outdir, xlsx_to_json_name(path))
        with open(out_name, "w", encoding="utf-8") as f:
            json.dump([device], f, ensure_ascii=False)

        n_groups = len(device["elements"])
        n_regs   = sum(len(el["elements"]) for el in device["elements"])
        print(f"  → {out_name}  ({n_groups} Gruppen, {n_regs} Register)")
        ok += 1

    print(f"\nFertig: {ok}/{len(paths)} Dateien verarbeitet.")


if __name__ == "__main__":
    main()
